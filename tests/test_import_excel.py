"""Tests for import_excel — both formats, dedup, edge cases."""
import pytest
from datetime import datetime, time as dt_time
from openpyxl import Workbook

import storage
from import_excel import import_from_file, _normalise_time


# ── _normalise_time ───────────────────────────────────────────────────────────

def test_normalise_time_object():
    assert _normalise_time(dt_time(9, 30)) == "09:30"


def test_normalise_datetime_object():
    assert _normalise_time(datetime(2026, 5, 10, 17, 0)) == "17:00"


def test_normalise_float():
    # 9.30 means 09:30 in Finnish spreadsheet convention
    assert _normalise_time(9.30) == "09:30"


def test_normalise_float_string():
    assert _normalise_time("16.45") == "16:45"


def test_normalise_none():
    assert _normalise_time(None) is None


def test_normalise_invalid_string():
    assert _normalise_time("abc") is None


# ── import helpers ────────────────────────────────────────────────────────────

def _make_simple_xlsx(path, rows):
    """Create a Simple-format xlsx: date | time_in | time_out."""
    wb = Workbook(); ws = wb.active
    for date, t_in, t_out in rows:
        ws.append([date, t_in, t_out])
    wb.save(path)


def _make_eas_xlsx(path, rows):
    """Create an EAS-format xlsx: date | job_shift | time_in | time_out."""
    wb = Workbook(); ws = wb.active
    for date, shift, t_in, t_out in rows:
        ws.append([date, shift, t_in, t_out])
    wb.save(path)


# ── Simple format ─────────────────────────────────────────────────────────────

def test_import_simple_one_row(tmp_path, tmp_storage):
    p = str(tmp_path / "t.xlsx")
    _make_simple_xlsx(p, [(datetime(2026, 5, 10), dt_time(9, 0), dt_time(17, 0))])
    imported, skipped = import_from_file(p)
    assert imported == 1
    assert skipped == 0


def test_import_simple_multiple_rows(tmp_path, tmp_storage):
    p = str(tmp_path / "t.xlsx")
    _make_simple_xlsx(p, [
        (datetime(2026, 5, 10), dt_time(9, 0), dt_time(17, 0)),
        (datetime(2026, 5, 11), dt_time(8, 0), dt_time(16, 0)),
    ])
    imported, skipped = import_from_file(p)
    assert imported == 2
    assert skipped == 0


def test_import_simple_entry_saved(tmp_path, tmp_storage):
    p = str(tmp_path / "t.xlsx")
    _make_simple_xlsx(p, [(datetime(2026, 5, 10), dt_time(9, 0), dt_time(17, 0))])
    import_from_file(p)
    entries = storage.load_month(2026, 5)
    assert len(entries) == 1
    assert entries[0].time_in  == "09:00"
    assert entries[0].time_out == "17:00"


# ── EAS format ────────────────────────────────────────────────────────────────

def test_import_eas_preserves_shift(tmp_path, tmp_storage):
    p = str(tmp_path / "t.xlsx")
    _make_eas_xlsx(p, [
        (datetime(2026, 5, 10), "Sales", dt_time(9, 0), dt_time(17, 0)),
    ])
    import_from_file(p)
    entries = storage.load_month(2026, 5)
    assert entries[0].job_shift == "Sales"


def test_import_eas_multiple_shifts(tmp_path, tmp_storage):
    p = str(tmp_path / "t.xlsx")
    _make_eas_xlsx(p, [
        (datetime(2026, 5, 10), "Sales",   dt_time(9,  0), dt_time(17, 0)),
        (datetime(2026, 5, 11), "Support", dt_time(10, 0), dt_time(18, 0)),
    ])
    imported, _ = import_from_file(p)
    assert imported == 2
    entries = storage.load_month(2026, 5)
    shifts = {e.job_shift for e in entries}
    assert shifts == {"Sales", "Support"}


# ── Duplicate detection ───────────────────────────────────────────────────────

def test_duplicate_on_second_import(tmp_path, tmp_storage):
    p = str(tmp_path / "t.xlsx")
    _make_simple_xlsx(p, [(datetime(2026, 5, 10), dt_time(9, 0), dt_time(17, 0))])
    import_from_file(p)
    imported, skipped = import_from_file(p)
    assert imported == 0
    assert skipped == 1


def test_no_duplicate_different_time(tmp_path, tmp_storage):
    p1 = str(tmp_path / "a.xlsx")
    p2 = str(tmp_path / "b.xlsx")
    _make_simple_xlsx(p1, [(datetime(2026, 5, 10), dt_time(9, 0), dt_time(17, 0))])
    _make_simple_xlsx(p2, [(datetime(2026, 5, 10), dt_time(10, 0), dt_time(18, 0))])
    import_from_file(p1)
    imported, skipped = import_from_file(p2)
    assert imported == 1
    assert skipped == 0


# ── Edge cases ────────────────────────────────────────────────────────────────

def test_missing_time_in_skipped(tmp_path, tmp_storage):
    """Rows with no time_in must be skipped without crashing."""
    wb = Workbook(); ws = wb.active
    ws.append([datetime(2026, 5, 10), None, None])   # no times
    p = str(tmp_path / "t.xlsx")
    wb.save(p)
    imported, _ = import_from_file(p)
    assert imported == 0


def test_non_date_rows_ignored(tmp_path, tmp_storage):
    """Header rows and text rows must not crash the importer."""
    wb = Workbook(); ws = wb.active
    ws.append(["Date", "Time In", "Time Out"])          # header row
    ws.append([datetime(2026, 5, 10), dt_time(9, 0), dt_time(17, 0)])
    p = str(tmp_path / "t.xlsx")
    wb.save(p)
    imported, _ = import_from_file(p)
    assert imported == 1
