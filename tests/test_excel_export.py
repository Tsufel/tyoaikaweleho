"""Tests for excel_export — _time_fraction, Simple and Full formats."""
import pytest
from openpyxl import load_workbook

import storage
from storage import WorkEntry
from excel_export import _time_fraction, export_month


# ── _time_fraction ────────────────────────────────────────────────────────────

def test_time_fraction_midnight():
    assert _time_fraction("00:00") == pytest.approx(0.0)


def test_time_fraction_noon():
    assert _time_fraction("12:00") == pytest.approx(0.5)


def test_time_fraction_nine_thirty():
    assert _time_fraction("09:30") == pytest.approx(9.5 / 24)


def test_time_fraction_end_of_day():
    assert _time_fraction("23:59") == pytest.approx((23 * 60 + 59) / (24 * 60))


# ── helpers ───────────────────────────────────────────────────────────────────

def _sample_entries():
    return [
        WorkEntry(id="1", date="2026-05-10", job_shift="Sales",
                  time_in="09:00", time_out="17:00"),
        WorkEntry(id="2", date="2026-05-11", job_shift="Support",
                  time_in="10:00", time_out="18:00"),
    ]


# ── Simple format ─────────────────────────────────────────────────────────────

def test_simple_headers(tmp_path):
    out = tmp_path / "out.xlsx"
    export_month(_sample_entries(), 2026, 5, str(out), fmt="Simple")
    ws = load_workbook(str(out)).active
    assert ws["A1"].value == "PVM"
    assert ws["C1"].value == "Start"
    assert ws["D1"].value == "End"
    assert ws["E1"].value == "Total (h)"


def test_simple_data_rows(tmp_path):
    out = tmp_path / "out.xlsx"
    export_month(_sample_entries(), 2026, 5, str(out), fmt="Simple")
    ws = load_workbook(str(out)).active
    # Row 2 should have a date value (openpyxl returns Python date objects)
    assert ws["A2"].value is not None


def test_simple_total_formula_present(tmp_path):
    out = tmp_path / "out.xlsx"
    export_month(_sample_entries(), 2026, 5, str(out), fmt="Simple")
    ws = load_workbook(str(out)).active
    # Find any cell containing a SUM formula
    found_sum = any(
        str(ws.cell(row=r, column=c).value or "").startswith("=SUM")
        for r in range(1, ws.max_row + 1)
        for c in range(1, ws.max_column + 1)
    )
    assert found_sum


def test_simple_include_pay_adds_rows(tmp_path):
    out = tmp_path / "out.xlsx"
    export_month(_sample_entries(), 2026, 5, str(out),
                 fmt="Simple", include_pay=True, pay_rate=22.5)
    ws = load_workbook(str(out)).active
    # Find "Rate" and "Earned" labels somewhere in column G
    col_g = [ws.cell(row=r, column=7).value for r in range(1, ws.max_row + 1)]
    assert "Rate"   in col_g
    assert "Earned" in col_g


def test_simple_no_pay_no_rate_row(tmp_path):
    out = tmp_path / "out.xlsx"
    export_month(_sample_entries(), 2026, 5, str(out),
                 fmt="Simple", include_pay=False)
    ws = load_workbook(str(out)).active
    col_g = [ws.cell(row=r, column=7).value for r in range(1, ws.max_row + 1)]
    assert "Rate" not in col_g


def test_simple_empty_entries(tmp_path):
    """Empty entry list must not raise and must still produce a valid file."""
    out = tmp_path / "out.xlsx"
    export_month([], 2026, 5, str(out), fmt="Simple")
    wb = load_workbook(str(out))
    assert wb.active is not None


# ── Full format ───────────────────────────────────────────────────────────────

def test_full_sheet_name(tmp_path):
    out = tmp_path / "out.xlsx"
    export_month(_sample_entries(), 2026, 5, str(out), fmt="Full")
    wb = load_workbook(str(out))
    assert wb.active.title == "MONTHLY TIMESHEET"


def test_full_column_headers(tmp_path):
    out = tmp_path / "out.xlsx"
    export_month(_sample_entries(), 2026, 5, str(out), fmt="Full")
    ws = load_workbook(str(out)).active
    # Headers are in row 11
    assert ws["A11"].value == "DATE"
    assert ws["C11"].value == "TIME IN"
    assert ws["D11"].value == "TIME OUT"


def test_full_empty_entries(tmp_path):
    out = tmp_path / "out.xlsx"
    export_month([], 2026, 5, str(out), fmt="Full")
    assert load_workbook(str(out)).active is not None
