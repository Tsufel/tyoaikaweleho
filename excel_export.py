from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from storage import WorkEntry
from utils import is_overnight


# ── shared helpers ─────────────────────────────────────────────────────────────

def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _time_fraction(hhmm: str) -> float:
    """Convert HH:MM string to Excel time fraction."""
    parts = hhmm.split(":")
    h, m = int(parts[0]), int(parts[1])
    return (h * 60 + m) / (24 * 60)


# ── public entry point ─────────────────────────────────────────────────────────

def export_month(entries: list[WorkEntry], year: int, month: int, save_path: str,
                 fmt: str = "Simple", include_pay: bool = False, pay_rate: float = 20.0):
    """Export *entries* to an Excel file at *save_path*.

    fmt:         "Simple" → clean minimal table (default)
                 "Full" → original elaborate EAS layout
    include_pay: when True and fmt=="Simple", append Rate + Earned rows
    pay_rate:    hourly rate used for the earnings calculation
    """
    if fmt == "Full":
        _export_full(entries, year, month, save_path, pay_rate)
    else:
        _export_simple(entries, year, month, save_path, include_pay, pay_rate)


# ── Simple layout ──────────────────────────────────────────────────────────────

def _export_simple(entries: list[WorkEntry], year: int, month: int,
                   save_path: str, include_pay: bool, pay_rate: float):
    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 12

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    # Row 1 — headers
    ws["A1"] = "PVM"
    ws["A1"].font = bold
    ws["C1"] = "Start"
    ws["C1"].font = bold
    ws["D1"] = "End"
    ws["D1"].font = bold
    ws["E1"] = "Total (h)"
    ws["E1"].font = bold

    # Data rows — sorted by (date, time_in), starting at row 2
    sorted_entries = sorted(entries, key=lambda e: (e.date, e.time_in))

    for i, entry in enumerate(sorted_entries):
        r = i + 2
        ws.row_dimensions[r].height = 16

        d = date.fromisoformat(entry.date)
        ws[f"A{r}"] = d
        ws[f"A{r}"].number_format = "DD.MM.YYYY"

        if entry.time_in:
            ws[f"C{r}"] = _time_fraction(entry.time_in)
            ws[f"C{r}"].number_format = "HH:MM"
            ws[f"C{r}"].alignment = center

        if entry.time_out:
            frac = _time_fraction(entry.time_out)
            if entry.time_in and is_overnight(entry.time_in, entry.time_out):
                frac += 1.0
            ws[f"D{r}"] = frac
            ws[f"D{r}"].number_format = "HH:MM"
            ws[f"D{r}"].alignment = center

            ws[f"E{r}"] = f"=D{r}-C{r}"
            ws[f"E{r}"].number_format = "[H]:MM"
            ws[f"E{r}"].alignment = center

    # Summary — Total label + SUM formula in column G
    last_data_row = len(sorted_entries) + 1  # row number of last data row
    total_label_row = last_data_row + 2
    total_sum_row   = last_data_row + 3

    ws[f"G{total_label_row}"] = "Total"
    ws[f"G{total_label_row}"].font = bold

    if sorted_entries:
        ws[f"G{total_sum_row}"] = f"=SUM(E2:E{last_data_row})"
    else:
        ws[f"G{total_sum_row}"] = 0
    ws[f"G{total_sum_row}"].number_format = "[H]:MM"
    ws[f"G{total_sum_row}"].font = bold

    # Optional pay section
    if include_pay:
        rate_row   = total_sum_row + 2
        earned_row = total_sum_row + 3

        ws[f"G{rate_row}"] = "Rate"
        ws[f"G{rate_row}"].font = bold
        ws[f"H{rate_row}"] = pay_rate
        ws[f"H{rate_row}"].number_format = "#,##0.00"

        ws[f"G{earned_row}"] = "Earned"
        ws[f"G{earned_row}"].font = bold
        ws[f"H{earned_row}"] = f"=G{total_sum_row}*24*H{rate_row}"
        ws[f"H{earned_row}"].number_format = '#,##0.00\\ "€"'
        ws[f"H{earned_row}"].font = Font(bold=True)

    wb.save(save_path)


# ── Full layout ──────────────────────────────────────────────────────────

def _export_full(entries: list[WorkEntry], year: int, month: int,
                 save_path: str, pay_rate: float):
    import calendar
    wb = Workbook()
    ws = wb.active
    ws.title = "MONTHLY TIMESHEET"

    # column widths
    widths = {"A": 14, "B": 16, "C": 10, "D": 10, "E": 10,
              "F": 6, "G": 12, "H": 14, "I": 10, "J": 8,
              "K": 8, "L": 8, "M": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    title_font  = Font(name="Arial", bold=True, size=14)
    header_font = Font(name="Arial", bold=True, size=10)
    label_font  = Font(name="Arial", bold=True, size=10)
    normal_font = Font(name="Arial", size=10)
    gray_fill   = PatternFill("solid", start_color="D9D9D9")
    center      = Alignment(horizontal="center", vertical="center")
    left        = Alignment(horizontal="left",   vertical="center")

    # Row 1: Title
    ws.merge_cells("A1:M1")
    ws.row_dimensions[1].height = 28
    c = ws["A1"]
    c.value = "MONTHLY TIMESHEET  EAS"
    c.font  = Font(name="Arial", bold=True, size=16)
    c.alignment = center
    c.fill  = gray_fill

    # Row 3: Employee
    ws["A3"] = "EMPLOYEE:"; ws["A3"].font = label_font
    ws["D3"] = "SIGNATUR:"; ws["D3"].font = label_font
    ws["G3"] = "DATE:";     ws["G3"].font = label_font
    ws["H3"] = datetime.today().strftime("%d.%m.%Y"); ws["H3"].font = normal_font

    # Row 5: Manager
    ws["A5"] = "MANAGER:";  ws["A5"].font = label_font
    ws["D5"] = "SIGNATUR:"; ws["D5"].font = label_font
    ws["G5"] = "DATE:";     ws["G5"].font = label_font

    # Row 7/8: Month and pay rate
    ws["D7"] = "MONTH / YEAR"; ws["D7"].font = label_font
    month_name = date(year, month, 1).strftime("%B %Y").upper()
    ws["E7"] = month_name; ws["E7"].font = normal_font
    ws["G7"] = "STANDARD";  ws["G7"].font = label_font
    ws["G8"] = "PAY RATE:"; ws["G8"].font = label_font
    ws["H8"] = pay_rate
    ws["H8"].font = Font(name="Arial", bold=True, size=10)
    ws["H8"].number_format = "#,##0.00"

    # Row 10: TOTAL label
    ws["E10"] = "TOTAL"; ws["E10"].font = label_font; ws["E10"].alignment = center

    # Row 11: Column headers
    ws.row_dimensions[11].height = 18
    headers = {"A": "DATE", "B": "JOB / SHIFT", "C": "TIME IN",
               "D": "TIME OUT", "E": "(HOURS)"}
    for col, text in headers.items():
        cell = ws[f"{col}11"]
        cell.value = text; cell.font = header_font
        cell.fill  = gray_fill; cell.alignment = center
        cell.border = _thin_border()

    # Group entries by ISO week
    by_week: dict[int, list[WorkEntry]] = {}
    for e in sorted(entries, key=lambda x: (x.date, x.time_in)):
        d  = date.fromisoformat(e.date)
        wk = d.isocalendar()[1]
        by_week.setdefault(wk, []).append(e)

    data_rows: list[int] = []
    current_row = 13

    for week_num in sorted(by_week.keys()):
        for entry in by_week[week_num]:
            r = current_row
            data_rows.append(r)
            ws.row_dimensions[r].height = 16

            d = date.fromisoformat(entry.date)
            ws[f"A{r}"] = d.strftime("%a %d %b %Y")
            ws[f"A{r}"].font = normal_font; ws[f"A{r}"].alignment = left

            ws[f"B{r}"] = entry.job_shift
            ws[f"B{r}"].font = normal_font; ws[f"B{r}"].alignment = center

            if entry.time_in:
                ws[f"C{r}"] = _time_fraction(entry.time_in)
                ws[f"C{r}"].number_format = "HH:MM"
                ws[f"C{r}"].alignment = center; ws[f"C{r}"].font = normal_font

            if entry.time_out:
                frac = _time_fraction(entry.time_out)
                if entry.time_in and is_overnight(entry.time_in, entry.time_out):
                    frac += 1.0
                ws[f"D{r}"] = frac
                ws[f"D{r}"].number_format = "HH:MM"
                ws[f"D{r}"].alignment = center; ws[f"D{r}"].font = normal_font

                ws[f"E{r}"] = f"=D{r}-C{r}"
                ws[f"E{r}"].number_format = "[H]:MM"
                ws[f"E{r}"].alignment = center; ws[f"E{r}"].font = normal_font

            for col in "ABCDE":
                ws[f"{col}{r}"].border = _thin_border()

            current_row += 1
        current_row += 2  # blank spacer rows between weeks

    # Summary section
    summary_start = current_row + 1

    ws[f"H{summary_start}"] = "HOURS"
    ws[f"H{summary_start}"].font = label_font
    ws[f"H{summary_start}"].alignment = center

    ws[f"H{summary_start+1}"] = "THIS MONTH"
    ws[f"H{summary_start+1}"].font = label_font
    ws[f"H{summary_start+1}"].alignment = center

    if data_rows:
        hours_refs   = "+".join([f"E{r}" for r in data_rows])
        total_formula = f"={hours_refs}"
    else:
        total_formula = "=0"

    ws[f"I{summary_start+1}"] = total_formula
    ws[f"I{summary_start+1}"].number_format = "[H]:MM"
    ws[f"I{summary_start+1}"].font = Font(name="Arial", bold=True, size=10)
    ws[f"I{summary_start+1}"].alignment = center
    ws[f"I{summary_start+1}"].border = _thin_border()

    ws[f"E{summary_start+3}"] = "RATE"; ws[f"E{summary_start+3}"].font = label_font

    subtotal_row = summary_start + 5
    ws[f"D{subtotal_row}"] = "SUB-TOTAL"; ws[f"D{subtotal_row}"].font = label_font
    ws[f"E{subtotal_row}"] = f"=I{summary_start+1}*24*H8"
    ws[f"E{subtotal_row}"].number_format = "#,##0.00"
    ws[f"E{subtotal_row}"].font = Font(name="Arial", bold=True, size=10)
    ws[f"E{subtotal_row}"].border = _thin_border()

    total_row = subtotal_row + 2
    ws[f"L{total_row}"] = "TOTAL"; ws[f"L{total_row}"].font = label_font
    ws[f"M{total_row}"] = f"=E{subtotal_row}"
    ws[f"M{total_row}"].number_format = "#,##0.00"
    ws[f"M{total_row}"].font = Font(name="Arial", bold=True, size=10)
    ws[f"M{total_row}"].border = _thin_border()

    wb.save(save_path)
