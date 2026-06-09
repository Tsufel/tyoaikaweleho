from datetime import datetime, time as dt_time
from openpyxl import load_workbook

import storage


def _normalise_time(val) -> str | None:
    """Convert any time representation to HH:MM string."""
    if val is None:
        return None
    if isinstance(val, dt_time):
        return f"{val.hour:02d}:{val.minute:02d}"
    if isinstance(val, datetime):
        return f"{val.hour:02d}:{val.minute:02d}"
    # Text or float like 16.45 or "16.45"
    try:
        f = float(str(val).replace(",", "."))
        hours = int(f)
        minutes = round((f - hours) * 100)
        if minutes >= 60:
            minutes = 59
        return f"{hours:02d}:{minutes:02d}"
    except (ValueError, TypeError):
        return None


def _detect_format(ws) -> str:
    """Return 'simple' (date|time_in|time_out) or 'eas' (date|shift|time_in|time_out)."""
    for row in ws.iter_rows(min_row=1, values_only=True):
        if isinstance(row[0], datetime) and len(row) > 1:
            return "simple" if isinstance(row[1], (dt_time, datetime)) else "eas"
    return "eas"


def import_from_file(path: str) -> tuple[int, int]:
    """Parse a timesheet Excel file and insert entries into storage.

    Handles two formats automatically:
    - EAS format:    date | job_shift | time_in | time_out
    - Simple format: date | time_in   | time_out | (anything)

    Returns (imported_count, skipped_count).
    Skips duplicates (same date + time_in already in storage).
    """
    wb = load_workbook(path, data_only=True)

    existing = storage.load_all_entries()
    existing_keys = {(e.date, e.time_in) for e in existing}

    imported = 0
    skipped = 0

    for ws in wb.worksheets:
        fmt = _detect_format(ws)
        last_date_str: str | None = None

        for row in ws.iter_rows(min_row=1, values_only=True):
            date_val = row[0] if row else None

            if isinstance(date_val, datetime):
                last_date_str = date_val.strftime("%Y-%m-%d")
                date_str = last_date_str
            elif date_val is None and last_date_str is not None:
                date_str = last_date_str
            else:
                last_date_str = None
                continue

            if fmt == "simple":
                job_shift = "GPSR"
                time_in  = _normalise_time(row[1]) if len(row) > 1 else None
                time_out = _normalise_time(row[2]) if len(row) > 2 else None
            else:  # eas
                raw_shift = row[1] if len(row) > 1 else None
                job_shift = str(raw_shift).strip() if raw_shift is not None else ""
                if not job_shift or job_shift.lower() == "none":
                    job_shift = "GPSR"
                time_in  = _normalise_time(row[2]) if len(row) > 2 else None
                time_out = _normalise_time(row[3]) if len(row) > 3 else None

            if not time_in:
                continue

            key = (date_str, time_in)
            if key in existing_keys:
                skipped += 1
                continue

            entry = storage.WorkEntry(
                id=storage.new_entry_id(),
                date=date_str,
                job_shift=job_shift,
                time_in=time_in,
                time_out=time_out or "",
            )
            storage.save_entry(entry)
            existing_keys.add(key)
            imported += 1

    return imported, skipped
